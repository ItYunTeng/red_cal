# from openpyxl import Workbook
#
# # 创建一个新的工作簿
# wb = Workbook()
#
# # 选择要添加数据的工作表，如果不存在则创建
# ws = wb.active
#
# # 一对多的数据
# one_to_many_data = {
#     "A1": ["数据1", "数据2", "数据3"],
#     "B1": ["数据4", "数据5"],
#     "C1": ["数据6"]
# }
#
# # 向工作表添加一对多的数据行
# for row, data in one_to_many_data.items():
#     ws[row] = data
#
# # 保存工作簿
# wb.save("output.xlsx")

import pandas as pd

# 示例一对多数据
data = {
    '主题': ['主题1', '主题1', '主题2', '主题2', '主题2'],
    '细节': ['细节1', '细节2', '细节1', '细节2', '细节3'],
    '描述': ['描述1', '描述1', '描述2', '描述2', '描述2']
}

# 创建DataFrame
df = pd.DataFrame(data)

# 设置主题为索引进行一对多展开
df_exploded = df.set_index('主题').stack().reset_index()
df_exploded.columns = ['主题', '细节', '描述']

# 将DataFrame写入Excel文件
with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
    df_exploded.to_excel(writer, index=False)

# import pandas as pd
#pip install pandas openpyxl
# # 读取Excel文件
# df = pd.read_excel('example.xlsx')
#
# # 合并前n列
# n = 2  # 假设我们要合并前两列
# merged_columns = df.iloc[:, :n].merge(df.iloc[:, 0], left_index=True, right_index=True)
#
# # 合并前n行
# n = 3  # 假设我们要合并前三行
# merged_rows = df.iloc[:n, :].merge(df.iloc[0, :], left_index=True, right_index=True)
#
# # 输出结果
# print(merged_columns)
# print(merged_rows)

# import pandas as pd
#
# # 读取Excel文件
# df = pd.read_excel('data.xlsx')
#
# # 按照某列进行分组，例如按照'A'列分组
# grouped = df.groupby('A')
#
# # 计算每个分组的统计信息，例如计数和平均值
# grouped_info = grouped.agg({'B': 'count', 'C': 'mean'})
#
# # 将结果输出到新的Excel文件
# grouped_info.to_excel('grouped_data.xlsx', index=False)


from openpyxl import load_workbook

# 加载 Excel 文件
workbook = load_workbook('example.xlsx')
sheet = workbook.active

# 遍历所有单元格，去除换行符和头尾空格
for row in sheet.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            # 去掉换行符和两端空格
            cell.value = cell.value.replace('\n', '').strip()

# 保存修改后的文件
workbook.save('example_cleaned.xlsx')
